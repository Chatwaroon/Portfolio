# wikitoexcel.py
# pip install wikipedia

import wikipedia
import random
wikipedia.set_lang('th')

keyword = ['toyota yaris','honda accord','suzuki swift']
allresult = []
allprice = []

for k in keyword:
	result = wikipedia.summary(k)
	print(result)
	allresult.append(result)
	allprice.append(random.randint(500,1000))
	print('-------------')

#####################

from openpyxl import Workbook

excelfile = Workbook()
sheet = excelfile.active

sheet['A1'].value = 'Keyword'
sheet['B1'].value = 'Detail'
sheet['C1'].value = 'Price'


for i,pd in enumerate(zip(keyword,allresult,allprice)):
	sheet.cell(row=i+2,column=1).value = pd[0]
	sheet.cell(row=i+2,column=2).value = pd[1]
	sheet.cell(row=i+2,column=3).value = pd[2]
# for i in range(2,50):
# 	sheet.cell(row=i,column=1).value = i

excelfile.save('product.xlsx')